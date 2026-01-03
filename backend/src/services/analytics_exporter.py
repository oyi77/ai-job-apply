"""
Export functionality for analytics reports.

Supports PDF, CSV, and Excel formats.
"""

import io
import csv
from typing import List, Dict, Any
from datetime import datetime
from loguru import logger


class AnalyticsExporter:
    """Export analytics data to various formats."""
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "analytics_report") -> bytes:
        """
        Export analytics data to CSV format.
        
        Args:
            data: List of dictionaries containing analytics data
            filename: Base filename for the export
            
        Returns:
            CSV data as bytes
        """
        try:
            output = io.StringIO()
            
            if not data:
                return b""
            
            # Get headers from first row
            headers = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=headers)
            
            writer.writeheader()
            writer.writerows(data)
            
            # Convert to bytes
            csv_data = output.getvalue().encode('utf-8')
            output.close()
            
            return csv_data
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}", exc_info=True)
            raise
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = "analytics_report") -> bytes:
        """
        Export analytics data to Excel format.
        
        Requires openpyxl library.
        """
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                logger.warning("openpyxl not installed, falling back to CSV")
                return AnalyticsExporter.export_to_csv(data, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analytics Report"
            
            if not data:
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
            
            # Add headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            raise
    
    @staticmethod
    def export_to_pdf(data: Dict[str, Any], filename: str = "analytics_report") -> bytes:
        """
        Export analytics data to PDF format.
        
        Requires reportlab library.
        """
        try:
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib import colors
            except ImportError:
                logger.warning("reportlab not installed, PDF export not available")
                raise ImportError("reportlab library required for PDF export")
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
            )
            title = Paragraph("Analytics Report", title_style)
            elements.append(title)
            
            # Timestamp
            timestamp = Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            )
            elements.append(timestamp)
            elements.append(Spacer(1, 0.3 * inch))
            
            # Add sections
            for section_name, section_data in data.items():
                # Section header
                section_title = Paragraph(
                    section_name.replace('_', ' ').title(),
                    styles['Heading2']
                )
                elements.append(section_title)
                elements.append(Spacer(1, 0.2 * inch))
                
                # Section data as table
                if isinstance(section_data, dict):
                    table_data = [[k.replace('_', ' ').title(), str(v)] for k, v in section_data.items()]
                    table = Table(table_data, colWidths=[3 * inch, 3 * inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(table)
                
                elements.append(Spacer(1, 0.3 * inch))
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}", exc_info=True)
            raise

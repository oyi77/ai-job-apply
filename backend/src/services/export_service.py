"""Export service implementation for generating reports in PDF, CSV, and Excel formats."""

import csv
import io
from typing import List, Optional, Dict, Any
from datetime import datetime
from pathlib import Path
from loguru import logger

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not available, PDF export will be limited")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available, Excel export will be limited")

from ..core.export_service import ExportService


class MultiFormatExportService(ExportService):
    """Export service supporting PDF, CSV, and Excel formats."""
    
    def __init__(self):
        """Initialize the export service."""
        self._logger = logger.bind(module="ExportService")
        self._supported_formats = ["csv", "excel", "xlsx"]
        if REPORTLAB_AVAILABLE:
            self._supported_formats.append("pdf")
    
    async def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats."""
        return self._supported_formats.copy()
    
    async def export_applications(
        self,
        applications: List[Dict[str, Any]],
        format: str = "csv",
        user_id: Optional[str] = None
    ) -> bytes:
        """Export job applications in the specified format."""
        if not applications:
            raise ValueError("No applications to export")
        
        format_lower = format.lower()
        
        if format_lower == "pdf":
            return await self._export_applications_pdf(applications)
        elif format_lower in ["excel", "xlsx"]:
            return await self._export_applications_excel(applications)
        elif format_lower == "csv":
            return await self._export_applications_csv(applications)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_resumes(
        self,
        resumes: List[Dict[str, Any]],
        format: str = "csv",
        user_id: Optional[str] = None
    ) -> bytes:
        """Export resumes in the specified format."""
        if not resumes:
            raise ValueError("No resumes to export")
        
        format_lower = format.lower()
        
        if format_lower == "pdf":
            return await self._export_resumes_pdf(resumes)
        elif format_lower in ["excel", "xlsx"]:
            return await self._export_resumes_excel(resumes)
        elif format_lower == "csv":
            return await self._export_resumes_csv(resumes)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_cover_letters(
        self,
        cover_letters: List[Dict[str, Any]],
        format: str = "pdf",
        user_id: Optional[str] = None
    ) -> bytes:
        """Export cover letters in the specified format."""
        if not cover_letters:
            raise ValueError("No cover letters to export")
        
        format_lower = format.lower()
        
        if format_lower == "pdf":
            return await self._export_cover_letters_pdf(cover_letters)
        elif format_lower in ["excel", "xlsx"]:
            return await self._export_cover_letters_excel(cover_letters)
        elif format_lower == "csv":
            return await self._export_cover_letters_csv(cover_letters)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    async def export_analytics(
        self,
        analytics_data: Dict[str, Any],
        format: str = "pdf",
        user_id: Optional[str] = None
    ) -> bytes:
        """Export analytics data in the specified format."""
        if not analytics_data:
            raise ValueError("No analytics data to export")
        
        format_lower = format.lower()
        
        if format_lower == "pdf":
            return await self._export_analytics_pdf(analytics_data)
        elif format_lower in ["excel", "xlsx"]:
            return await self._export_analytics_excel(analytics_data)
        elif format_lower == "csv":
            return await self._export_analytics_csv(analytics_data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    # PDF Export Methods
    def _export_applications_pdf(self, applications: List[Dict[str, Any]]) -> bytes:
        """Export applications to PDF."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("PDF export requires reportlab library")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Job Applications Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_text = f"Total Applications: {len(applications)}<br/>"
        summary_text += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Job Title', 'Company', 'Status', 'Applied Date', 'Notes']]
        for app in applications:
            applied_date = ""
            if app.get('applied_date'):
                try:
                    if isinstance(app['applied_date'], str):
                        dt = datetime.fromisoformat(app['applied_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['applied_date']
                    applied_date = dt.strftime('%Y-%m-%d')
                except:
                    applied_date = str(app.get('applied_date', ''))
            
            notes = str(app.get('notes', ''))[:50] + '...' if len(str(app.get('notes', ''))) > 50 else str(app.get('notes', ''))
            
            table_data.append([
                app.get('job_title', ''),
                app.get('company', ''),
                app.get('status', ''),
                applied_date,
                notes
            ])
        
        # Create table
        table = Table(table_data, colWidths=[2*inch, 2*inch, 1.2*inch, 1*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_resumes_pdf(self, resumes: List[Dict[str, Any]]) -> bytes:
        """Export resumes to PDF."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("PDF export requires reportlab library")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Resumes Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_text = f"Total Resumes: {len(resumes)}<br/>"
        summary_text += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Name', 'File Type', 'Skills', 'Experience', 'Default']]
        for resume in resumes:
            skills = ', '.join(resume.get('skills', [])[:5]) if resume.get('skills') else 'N/A'
            if len(skills) > 50:
                skills = skills[:50] + '...'
            
            table_data.append([
                resume.get('name', ''),
                resume.get('file_type', ''),
                skills,
                f"{resume.get('experience_years', 0)} years" if resume.get('experience_years') else 'N/A',
                'Yes' if resume.get('is_default') else 'No'
            ])
        
        table = Table(table_data, colWidths=[2.5*inch, 1*inch, 2.5*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_cover_letters_pdf(self, cover_letters: List[Dict[str, Any]]) -> bytes:
        """Export cover letters to PDF."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("PDF export requires reportlab library")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Cover Letters Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Add each cover letter
        for idx, cl in enumerate(cover_letters):
            if idx > 0:
                story.append(PageBreak())
            
            # Header
            story.append(Paragraph(f"Cover Letter #{idx + 1}", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            # Details
            details = f"<b>Job Title:</b> {cl.get('job_title', 'N/A')}<br/>"
            details += f"<b>Company:</b> {cl.get('company_name', 'N/A')}<br/>"
            details += f"<b>Tone:</b> {cl.get('tone', 'N/A')}<br/>"
            details += f"<b>Word Count:</b> {cl.get('word_count', 0)}<br/>"
            details += f"<b>Generated:</b> {cl.get('generated_at', 'N/A')}<br/>"
            story.append(Paragraph(details, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Content
            content_style = ParagraphStyle(
                'ContentStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                leftIndent=0.5*inch,
                rightIndent=0.5*inch
            )
            content = cl.get('content', '')
            # Split content into paragraphs
            paragraphs = content.split('\n\n')
            for para in paragraphs:
                if para.strip():
                    story.append(Paragraph(para.strip(), content_style))
                    story.append(Spacer(1, 0.1*inch))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_analytics_pdf(self, analytics_data: Dict[str, Any]) -> bytes:
        """Export analytics to PDF."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("PDF export requires reportlab library")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Analytics Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Statistics
        stats = analytics_data.get('statistics', {})
        if stats:
            story.append(Paragraph("Statistics", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            stats_text = ""
            for key, value in stats.items():
                stats_text += f"<b>{key.replace('_', ' ').title()}:</b> {value}<br/>"
            story.append(Paragraph(stats_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Additional data
        for section, data in analytics_data.items():
            if section == 'statistics':
                continue
            
            story.append(Paragraph(section.replace('_', ' ').title(), styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            if isinstance(data, dict):
                data_text = ""
                for key, value in data.items():
                    data_text += f"<b>{key.replace('_', ' ').title()}:</b> {value}<br/>"
                story.append(Paragraph(data_text, styles['Normal']))
            elif isinstance(data, list):
                for item in data[:10]:  # Limit to first 10 items
                    story.append(Paragraph(str(item), styles['Normal']))
            else:
                story.append(Paragraph(str(data), styles['Normal']))
            
            story.append(Spacer(1, 0.2*inch))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    # CSV Export Methods
    def _export_applications_csv(self, applications: List[Dict[str, Any]]) -> bytes:
        """Export applications to CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(['ID', 'Job Title', 'Company', 'Status', 'Applied Date', 'Follow Up Date', 'Interview Date', 'Notes'])
        
        # Data rows
        for app in applications:
            applied_date = ""
            if app.get('applied_date'):
                try:
                    if isinstance(app['applied_date'], str):
                        dt = datetime.fromisoformat(app['applied_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['applied_date']
                    applied_date = dt.strftime('%Y-%m-%d')
                except:
                    applied_date = str(app.get('applied_date', ''))
            
            follow_up = ""
            if app.get('follow_up_date'):
                try:
                    if isinstance(app['follow_up_date'], str):
                        dt = datetime.fromisoformat(app['follow_up_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['follow_up_date']
                    follow_up = dt.strftime('%Y-%m-%d')
                except:
                    follow_up = str(app.get('follow_up_date', ''))
            
            interview = ""
            if app.get('interview_date'):
                try:
                    if isinstance(app['interview_date'], str):
                        dt = datetime.fromisoformat(app['interview_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['interview_date']
                    interview = dt.strftime('%Y-%m-%d')
                except:
                    interview = str(app.get('interview_date', ''))
            
            writer.writerow([
                app.get('id', ''),
                app.get('job_title', ''),
                app.get('company', ''),
                app.get('status', ''),
                applied_date,
                follow_up,
                interview,
                app.get('notes', '')
            ])
        
        return buffer.getvalue().encode('utf-8-sig')  # UTF-8 with BOM for Excel compatibility
    
    def _export_resumes_csv(self, resumes: List[Dict[str, Any]]) -> bytes:
        """Export resumes to CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(['ID', 'Name', 'File Type', 'Skills', 'Experience Years', 'Education', 'Certifications', 'Default'])
        
        # Data rows
        for resume in resumes:
            skills = ', '.join(resume.get('skills', [])) if resume.get('skills') else ''
            education = ', '.join(resume.get('education', [])) if resume.get('education') else ''
            certifications = ', '.join(resume.get('certifications', [])) if resume.get('certifications') else ''
            
            writer.writerow([
                resume.get('id', ''),
                resume.get('name', ''),
                resume.get('file_type', ''),
                skills,
                resume.get('experience_years', ''),
                education,
                certifications,
                'Yes' if resume.get('is_default') else 'No'
            ])
        
        return buffer.getvalue().encode('utf-8-sig')
    
    def _export_cover_letters_csv(self, cover_letters: List[Dict[str, Any]]) -> bytes:
        """Export cover letters to CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(['ID', 'Job Title', 'Company', 'Tone', 'Word Count', 'Generated At', 'Content Preview'])
        
        # Data rows
        for cl in cover_letters:
            content_preview = cl.get('content', '')[:100] + '...' if len(cl.get('content', '')) > 100 else cl.get('content', '')
            
            generated_at = ""
            if cl.get('generated_at'):
                try:
                    if isinstance(cl['generated_at'], str):
                        dt = datetime.fromisoformat(cl['generated_at'].replace('Z', '+00:00'))
                    else:
                        dt = cl['generated_at']
                    generated_at = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    generated_at = str(cl.get('generated_at', ''))
            
            writer.writerow([
                cl.get('id', ''),
                cl.get('job_title', ''),
                cl.get('company_name', ''),
                cl.get('tone', ''),
                cl.get('word_count', 0),
                generated_at,
                content_preview
            ])
        
        return buffer.getvalue().encode('utf-8-sig')
    
    def _export_analytics_csv(self, analytics_data: Dict[str, Any]) -> bytes:
        """Export analytics to CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Write statistics
        stats = analytics_data.get('statistics', {})
        if stats:
            writer.writerow(['Metric', 'Value'])
            for key, value in stats.items():
                writer.writerow([key.replace('_', ' ').title(), value])
            writer.writerow([])  # Empty row
        
        # Write other data
        for section, data in analytics_data.items():
            if section == 'statistics':
                continue
            
            writer.writerow([section.replace('_', ' ').title()])
            if isinstance(data, dict):
                writer.writerow(['Key', 'Value'])
                for key, value in data.items():
                    writer.writerow([key, value])
            elif isinstance(data, list):
                if data and isinstance(data[0], dict):
                    # Write headers from first dict
                    headers = list(data[0].keys())
                    writer.writerow(headers)
                    for item in data:
                        writer.writerow([item.get(h, '') for h in headers])
                else:
                    writer.writerow(['Value'])
                    for item in data:
                        writer.writerow([str(item)])
            writer.writerow([])  # Empty row
        
        return buffer.getvalue().encode('utf-8-sig')
    
    # Excel Export Methods
    def _export_applications_excel(self, applications: List[Dict[str, Any]]) -> bytes:
        """Export applications to Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel export requires openpyxl library")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Job Title', 'Company', 'Status', 'Applied Date', 'Follow Up Date', 'Interview Date', 'Notes']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, app in enumerate(applications, 2):
            applied_date = ""
            if app.get('applied_date'):
                try:
                    if isinstance(app['applied_date'], str):
                        dt = datetime.fromisoformat(app['applied_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['applied_date']
                    applied_date = dt
                except:
                    applied_date = str(app.get('applied_date', ''))
            
            follow_up = ""
            if app.get('follow_up_date'):
                try:
                    if isinstance(app['follow_up_date'], str):
                        dt = datetime.fromisoformat(app['follow_up_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['follow_up_date']
                    follow_up = dt
                except:
                    follow_up = str(app.get('follow_up_date', ''))
            
            interview = ""
            if app.get('interview_date'):
                try:
                    if isinstance(app['interview_date'], str):
                        dt = datetime.fromisoformat(app['interview_date'].replace('Z', '+00:00'))
                    else:
                        dt = app['interview_date']
                    interview = dt
                except:
                    interview = str(app.get('interview_date', ''))
            
            row_data = [
                app.get('id', ''),
                app.get('job_title', ''),
                app.get('company', ''),
                app.get('status', ''),
                applied_date,
                follow_up,
                interview,
                app.get('notes', '')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_resumes_excel(self, resumes: List[Dict[str, Any]]) -> bytes:
        """Export resumes to Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel export requires openpyxl library")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumes"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Name', 'File Type', 'Skills', 'Experience Years', 'Education', 'Certifications', 'Default']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, resume in enumerate(resumes, 2):
            skills = ', '.join(resume.get('skills', [])) if resume.get('skills') else ''
            education = ', '.join(resume.get('education', [])) if resume.get('education') else ''
            certifications = ', '.join(resume.get('certifications', [])) if resume.get('certifications') else ''
            
            row_data = [
                resume.get('id', ''),
                resume.get('name', ''),
                resume.get('file_type', ''),
                skills,
                resume.get('experience_years', ''),
                education,
                certifications,
                'Yes' if resume.get('is_default') else 'No'
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_cover_letters_excel(self, cover_letters: List[Dict[str, Any]]) -> bytes:
        """Export cover letters to Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel export requires openpyxl library")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cover Letters"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Job Title', 'Company', 'Tone', 'Word Count', 'Generated At', 'Content']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, cl in enumerate(cover_letters, 2):
            generated_at = ""
            if cl.get('generated_at'):
                try:
                    if isinstance(cl['generated_at'], str):
                        dt = datetime.fromisoformat(cl['generated_at'].replace('Z', '+00:00'))
                    else:
                        dt = cl['generated_at']
                    generated_at = dt
                except:
                    generated_at = str(cl.get('generated_at', ''))
            
            row_data = [
                cl.get('id', ''),
                cl.get('job_title', ''),
                cl.get('company_name', ''),
                cl.get('tone', ''),
                cl.get('word_count', 0),
                generated_at,
                cl.get('content', '')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                # Wrap text for content column
                if col_idx == 7:  # Content column
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            if col_idx == 7:  # Content column
                ws.column_dimensions[column_letter].width = 60
            else:
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _export_analytics_excel(self, analytics_data: Dict[str, Any]) -> bytes:
        """Export analytics to Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel export requires openpyxl library")
        
        wb = Workbook()
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Statistics sheet
        stats = analytics_data.get('statistics', {})
        if stats:
            ws_stats = wb.active
            ws_stats.title = "Statistics"
            
            # Headers
            ws_stats.cell(row=1, column=1, value="Metric").fill = header_fill
            ws_stats.cell(row=1, column=1).font = header_font
            ws_stats.cell(row=1, column=1).alignment = header_alignment
            ws_stats.cell(row=1, column=1).border = border
            
            ws_stats.cell(row=1, column=2, value="Value").fill = header_fill
            ws_stats.cell(row=1, column=2).font = header_font
            ws_stats.cell(row=1, column=2).alignment = header_alignment
            ws_stats.cell(row=1, column=2).border = border
            
            # Data
            for row_idx, (key, value) in enumerate(stats.items(), 2):
                ws_stats.cell(row=row_idx, column=1, value=key.replace('_', ' ').title()).border = border
                ws_stats.cell(row=row_idx, column=2, value=value).border = border
                if row_idx % 2 == 0:
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    ws_stats.cell(row=row_idx, column=1).fill = fill
                    ws_stats.cell(row=row_idx, column=2).fill = fill
            
            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 20
        
        # Other data sheets
        sheet_idx = 1
        for section, data in analytics_data.items():
            if section == 'statistics':
                continue
            
            if sheet_idx == 1 and not stats:
                ws = wb.active
            else:
                ws = wb.create_sheet(title=section.replace('_', ' ').title()[:31])
            
            ws.title = section.replace('_', ' ').title()[:31]
            
            if isinstance(data, dict):
                # Headers
                ws.cell(row=1, column=1, value="Key").fill = header_fill
                ws.cell(row=1, column=1).font = header_font
                ws.cell(row=1, column=1).alignment = header_alignment
                ws.cell(row=1, column=1).border = border
                
                ws.cell(row=1, column=2, value="Value").fill = header_fill
                ws.cell(row=1, column=2).font = header_font
                ws.cell(row=1, column=2).alignment = header_alignment
                ws.cell(row=1, column=2).border = border
                
                # Data
                for row_idx, (key, value) in enumerate(data.items(), 2):
                    ws.cell(row=row_idx, column=1, value=str(key)).border = border
                    ws.cell(row=row_idx, column=2, value=str(value)).border = border
                    if row_idx % 2 == 0:
                        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        ws.cell(row=row_idx, column=1).fill = fill
                        ws.cell(row=row_idx, column=2).fill = fill
                
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 50
                
            elif isinstance(data, list) and data:
                if isinstance(data[0], dict):
                    # Headers from first dict
                    headers = list(data[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # Data rows
                    for row_idx, item in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            value = item.get(header, '')
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = border
                            if row_idx % 2 == 0:
                                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    # Auto-adjust widths
                    for col_idx in range(1, len(headers) + 1):
                        column_letter = get_column_letter(col_idx)
                        max_length = 0
                        for row in ws[column_letter]:
                            try:
                                if len(str(row.value)) > max_length:
                                    max_length = len(str(row.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                else:
                    # Simple list
                    ws.cell(row=1, column=1, value="Value").fill = header_fill
                    ws.cell(row=1, column=1).font = header_font
                    ws.cell(row=1, column=1).alignment = header_alignment
                    ws.cell(row=1, column=1).border = border
                    
                    for row_idx, item in enumerate(data, 2):
                        cell = ws.cell(row=row_idx, column=1, value=str(item))
                        cell.border = border
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    ws.column_dimensions['A'].width = 50
            
            sheet_idx += 1
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
